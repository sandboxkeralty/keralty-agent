import io
import datetime
from googleapiclient.discovery import build
import google.auth
from typing import Dict, Any, List

_SHEETS_SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive',
]

_XLSX_MIME_TYPES = {
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'application/vnd.ms-excel',
}

def get_sheets_service(credentials=None):
    if credentials is None:
        credentials, _ = google.auth.default(scopes=_SHEETS_SCOPES)
    return build('sheets', 'v4', credentials=credentials)

def get_drive_service(credentials=None):
    if credentials is None:
        credentials, _ = google.auth.default(scopes=_SHEETS_SCOPES)
    return build('drive', 'v3', credentials=credentials)

def _is_raw_excel_file(file_id: str, credentials=None) -> bool:
    drive = get_drive_service(credentials)
    meta = drive.files().get(fileId=file_id, fields="mimeType").execute()
    return meta.get("mimeType") in _XLSX_MIME_TYPES

_MAX_DOWNLOAD_BYTES = 50 * 1024 * 1024


def _load_xlsx_workbook(file_id: str, credentials=None):
    from openpyxl import load_workbook
    drive = get_drive_service(credentials)
    # Cap size before downloading so a huge workbook can't OOM the container.
    meta = drive.files().get(fileId=file_id, fields="size").execute()
    size = int(meta.get("size") or 0)
    if size > _MAX_DOWNLOAD_BYTES:
        raise ValueError(f"Spreadsheet is too large to read ({size // (1024 * 1024)} MB; limit 50 MB).")
    content = drive.files().get_media(fileId=file_id).execute()
    return load_workbook(io.BytesIO(content), data_only=True, read_only=True)

def _json_safe(value):
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    return value

class SheetsService:
    @staticmethod
    def get_spreadsheet(spreadsheet_id: str, credentials=None) -> Dict[str, Any]:
        if _is_raw_excel_file(spreadsheet_id, credentials):
            wb = _load_xlsx_workbook(spreadsheet_id, credentials)
            sheets = [
                {"properties": {"sheetId": i, "title": name, "index": i}}
                for i, name in enumerate(wb.sheetnames)
            ]
            return {"sheets": sheets}
        service = get_sheets_service(credentials)
        return service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()

    @staticmethod
    def read_range(spreadsheet_id: str, range_name: str, credentials=None) -> List[List[Any]]:
        if _is_raw_excel_file(spreadsheet_id, credentials):
            wb = _load_xlsx_workbook(spreadsheet_id, credentials)
            if "!" in range_name:
                tab_name, cell_range = range_name.split("!", 1)
            else:
                tab_name, cell_range = range_name, None
            ws = wb[tab_name] if tab_name in wb.sheetnames else wb.active
            rows = ws[cell_range] if cell_range else ws.iter_rows()
            values = [[_json_safe(c.value) for c in row] for row in rows]
            while values and all(v == "" for v in values[-1]):
                values.pop()
            return values
        service = get_sheets_service(credentials)
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id, range=range_name).execute()
        return result.get('values', [])

    @staticmethod
    def create_spreadsheet(title: str, credentials=None) -> Dict[str, Any]:
        service = get_sheets_service(credentials)
        spreadsheet = service.spreadsheets().create(
            body={'properties': {'title': title}},
            fields='spreadsheetId,sheets.properties'
        ).execute()
        spreadsheet_id = spreadsheet.get('spreadsheetId')
        sheets = spreadsheet.get('sheets', [])
        first_sheet_title = sheets[0]['properties']['title'] if sheets else 'Sheet1'
        return {'spreadsheet_id': spreadsheet_id, 'first_sheet_title': first_sheet_title}

    @staticmethod
    def update_values(spreadsheet_id: str, range_name: str, values: List[List[Any]], credentials=None) -> bool:
        service = get_sheets_service(credentials)
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=range_name,
            valueInputOption='USER_ENTERED',
            body={'values': values}
        ).execute()
        return True

    @staticmethod
    def append_values(spreadsheet_id: str, range_name: str, values: List[List[Any]], credentials=None) -> Dict[str, Any]:
        service = get_sheets_service(credentials)
        result = service.spreadsheets().values().append(
            spreadsheetId=spreadsheet_id,
            range=range_name,
            valueInputOption='USER_ENTERED',
            insertDataOption='INSERT_ROWS',
            body={'values': values}
        ).execute()
        return result.get('updates', {})

    @staticmethod
    def _require_native_sheet(spreadsheet_id: str, credentials=None) -> None:
        # Tab management goes through spreadsheets().batchUpdate, which only
        # understands native Google Sheets IDs — a raw uploaded .xlsx/.xls
        # can be *read* (via openpyxl) but not structurally modified here.
        if _is_raw_excel_file(spreadsheet_id, credentials):
            raise ValueError(
                "This file is a raw Excel upload (.xlsx/.xls); tabs can only be "
                "added, renamed or deleted on native Google Sheets. Convert the "
                "file to Google Sheets first."
            )

    @staticmethod
    def _resolve_sheet_id(spreadsheet_id: str, tab_title: str, credentials=None) -> int:
        meta = SheetsService.get_spreadsheet(spreadsheet_id, credentials)
        for s in meta.get("sheets", []):
            if s["properties"]["title"] == tab_title:
                return s["properties"]["sheetId"]
        titles = [s["properties"]["title"] for s in meta.get("sheets", [])]
        raise ValueError(f"No tab named '{tab_title}' in this spreadsheet. Existing tabs: {titles}")

    @staticmethod
    def add_sheet(spreadsheet_id: str, title: str, credentials=None) -> Dict[str, Any]:
        SheetsService._require_native_sheet(spreadsheet_id, credentials)
        service = get_sheets_service(credentials)
        result = service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": [{"addSheet": {"properties": {"title": title}}}]},
        ).execute()
        props = result["replies"][0]["addSheet"]["properties"]
        return {"sheet_id": props["sheetId"], "title": props["title"]}

    @staticmethod
    def rename_sheet(spreadsheet_id: str, tab_title: str, new_title: str, credentials=None) -> None:
        SheetsService._require_native_sheet(spreadsheet_id, credentials)
        sheet_id = SheetsService._resolve_sheet_id(spreadsheet_id, tab_title, credentials)
        service = get_sheets_service(credentials)
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": [{"updateSheetProperties": {
                "properties": {"sheetId": sheet_id, "title": new_title},
                "fields": "title",
            }}]},
        ).execute()

    @staticmethod
    def delete_sheet(spreadsheet_id: str, tab_title: str, credentials=None) -> None:
        SheetsService._require_native_sheet(spreadsheet_id, credentials)
        sheet_id = SheetsService._resolve_sheet_id(spreadsheet_id, tab_title, credentials)
        service = get_sheets_service(credentials)
        # The API itself rejects deleting the last remaining tab — that error
        # is surfaced to the caller rather than pre-checked here.
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": [{"deleteSheet": {"sheetId": sheet_id}}]},
        ).execute()

    @staticmethod
    def share_spreadsheet(spreadsheet_id: str, email: str, credentials=None) -> None:
        service = get_drive_service(credentials)
        service.permissions().create(
            fileId=spreadsheet_id,
            body={'type': 'user', 'role': 'writer', 'emailAddress': email},
            fields='id',
            sendNotificationEmail=False,
        ).execute()
